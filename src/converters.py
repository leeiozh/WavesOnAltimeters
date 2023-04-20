import datetime as dt
import numpy as np
import geopy.distance as dist
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill


def calc_alt(length, height):
    """
    calculate an angle above the horizon
    :param length: distance on Earth (radius of spot)
    :param height: height of orbit above the Earth
    :return: angle
    """
    phi = length / 6371
    s = 6371 * np.sin(phi)
    alph = np.arctan2(s, 6371 * (1 - np.cos(phi)) + height)
    return phi + alph


def utc_to_sec(utc: str, noyear=True, year=2022) -> float:
    """
    Конвертер из utc в секунды с J2000 для даты вида hh-mm-ss-yyyy-mm-dd
    :param utc: время в utc
    :param noyear: для укороченных названий
    :param year: для укороченных названий
    :return: время в секундах с J2000
    """
    if utc[2] != '-' and utc[2] != ':' and utc[2] != '.':
        if noyear:
            return (dt.datetime(int(utc[0:4]), int(utc[5:7]), int(utc[8:10]), int(utc[11:13]), int(utc[14:16]),
                                0) - dt.datetime(2000, 1, 1, 0, 0, 0)).total_seconds()
        else:
            return (dt.datetime(year, int(utc[5:7]), int(utc[8:10]), int(utc[11:13]), int(utc[14:16]),
                                0) - dt.datetime(2000, 1, 1, 0, 0, 0)).total_seconds()
    else:
        if len(utc) > 16:
            return (dt.datetime(int(utc[6:10]), int(utc[3:5]), int(utc[:2]), int(utc[11:13]), int(utc[14:16]),
                                int(utc[17:19])) - dt.datetime(2000, 1, 1, 0, 0, 0)).total_seconds()
        else:
            return (dt.datetime(int(utc[6:10]), int(utc[3:5]), int(utc[:2]), int(utc[11:13]), int(utc[14:16]),
                                0) - dt.datetime(2000, 1, 1, 0, 0, 0)).total_seconds()


def sec_to_utc(sec: float) -> dt.datetime:
    """
    Конвертер из секунд с J2000 в utc
    :param sec: время в секундах с J2000
    :return:  время в utc
    """
    return dt.datetime(2000, 1, 1, 0, 0, 0) + dt.timedelta(seconds=sec)


def to_deg(inp: str) -> float:
    arr = inp.split()
    if len(arr) > 1:
        return int(arr[0]) + float(arr[1]) / 60
    else:
        return float(arr[0])


def r_to_latlon(r) -> (np.ndarray, np.ndarray):
    norm = np.linalg.norm(r)
    lat = np.arccos(r[2] / norm) / np.pi * 180
    lon = np.arctan2(r[1], r[0]) / np.pi * 180
    return lat, lon


def time_to_sf(time: np.ndarray) -> list:
    res = np.ndarray(time.shape, dtype=dt.datetime)
    for i in range(len(res)):
        res[i] = dt.datetime(2000, 1, 1, 0, 0, 0, tzinfo=dt.timezone.utc) + dt.timedelta(seconds=time[i])
    return res.tolist()


def calc_time(track, start, speed):
    res = np.ndarray(track.shape[0] - 1, dtype=dt.datetime)
    res[0] = start
    for i in range(res.shape[0] - 1):
        res[i + 1] = res[i] + dt.timedelta(
            hours=(dist.geodesic((track[i + 1].y, track[i + 1].x), (track[i].y, track[i].x)).nm / speed))
    return res


def prepare_sheet(start, nums):
    res = pd.DataFrame(columns=['date'])
    res['date'] = [(start + dt.timedelta(days=i)).strftime("%Y-%m-%d") for i in range(nums)]
    return res


def convert_list(name_sheet, sheet, list):
    st = 0
    while len(list[st]) == 0:
        st += 1
    curr_date = list[st][0]['date']
    k = 0
    for i in range(st, len(list)):
        if len(list[i]) != 0:
            if curr_date != list[i][0]['date']:
                curr_date = list[i][0]['date']
                k = 0
            for j in range(len(list[i])):
                k += 1
                sheet.loc[sheet['date'] == curr_date, ("flight" + str(k))] = list[i][j]['time']

    sheet.to_excel(name_sheet, index=False, engine='openpyxl')
    wb = openpyxl.load_workbook(name_sheet)
    ws = wb['Sheet1']
    alhpabet = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']
    for i in range(len(list)):

        if len(list[i]) != 0:
            if curr_date != list[i][0]['date']:
                curr_date = list[i][0]['date']
                d = (sheet.index[sheet['date'] == curr_date][0]) + 2
                k = 0
            for j in range(len(list[i])):
                k += 1
                ws[str(alhpabet[k] + str(d))].fill = PatternFill(patternType='solid', fgColor=list[i][j]['color'])

    wb.save(name_sheet)
